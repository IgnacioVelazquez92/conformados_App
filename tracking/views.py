import re
import uuid
from datetime import date, datetime, time, timedelta
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.models import User
from django.conf import settings
from django.core.cache import cache
from django.core.paginator import Paginator
from django.core.files.base import File
from django.core.files.storage import default_storage
from django.db.models import Count, Max, Q
from django.http import HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.db.models.functions import TruncDate

from .forms import CierreHojaForm, EvidenciaForm, ImportPdfForm, ImportSpreadsheetForm, LoginForm, NoEntregadoForm, UserCreateForm, UserDeleteForm, UserUpdateForm, ValidacionEvidenciaForm
from .models import Evidencia, EventoTrazabilidad, HojaRuta, IntentoAccesoPortal, Remito
from .services.authz import (
    can_close_hoja,
    can_grant_staff,
    can_import_pdf,
    can_manage_users,
    can_review_evidence,
    create_user_with_profile,
    delete_user_and_profile,
    get_or_create_profile,
    update_user_with_profile,
)
from .services.admin_ops import cerrar_hoja_ruta, validar_evidencia as validar_evidencia_service
from .services.conformados import registrar_evidencia, registrar_intento_acceso_portal, registrar_intento_no_entregado
from .services.email_alerts import send_public_access_alert
from .services.import_pdf import _validate_parsed_hoja, extract_oid_from_qr, extract_text_from_pdf, import_hoja_ruta_pdf, parse_hoja_ruta_pdf
from .services.import_tabular import import_tabular_file, parse_tabular_file

REMITO_MANUAL_PATTERN = re.compile(r"^\d{5}-\d{8}$")
REMITO_MANUAL_DIGITS_PATTERN = re.compile(r"^\d{13}$")
OID_PATTERN = re.compile(r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}")
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
PDF_EXTENSIONS = {".pdf"}
IMPORT_PREVIEW_SESSION_KEY = "import_preview_files"


def _normalize_code(value: str) -> str:
    return re.sub(r"\W+", "", (value or "").upper())


def _get_client_ip(request: HttpRequest) -> str:
    forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR", "")
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR", "") or "unknown"


def _check_rate_limit(*, key: str, limit: int, window_seconds: int) -> None:
    count = cache.get(key, 0)
    if count >= limit:
        minutes = max(1, window_seconds // 60)
        raise ValueError(f"Demasiados intentos. Espera {minutes} minutos y volve a probar.")
    if count == 0:
        cache.set(key, 1, window_seconds)
    else:
        cache.incr(key)


def _rate_limit_key(*, action: str, request: HttpRequest, canal: str, oid: str, remito_uid: str) -> str:
    ip = _normalize_code(_get_client_ip(request))
    return f"rate:{action}:{ip}:{canal}:{oid}:{_normalize_code(remito_uid)}"


def _evidencia_limits_context() -> dict[str, int]:
    return {
        "max_image_mb": settings.EVIDENCIA_MAX_IMAGE_SIZE_MB,
        "max_pdf_mb": settings.EVIDENCIA_MAX_PDF_SIZE_MB,
        "max_image_bytes": settings.EVIDENCIA_MAX_IMAGE_SIZE_MB * 1024 * 1024,
        "max_pdf_bytes": settings.EVIDENCIA_MAX_PDF_SIZE_MB * 1024 * 1024,
    }


def _parse_dashboard_date(value: str | None, default: date) -> date:
    parsed = parse_date(value or "") if value else None
    return parsed or default


def _date_bounds(day: date) -> tuple[datetime, datetime]:
    tz = timezone.get_current_timezone()
    start = timezone.make_aware(datetime.combine(day, time.min), tz)
    end = timezone.make_aware(datetime.combine(day, time.max), tz)
    return start, end


def _build_day_series(start_day: date, end_day: date, points: dict[str, int]) -> list[int]:
    series: list[int] = []
    current_day = start_day
    while current_day <= end_day:
        series.append(int(points.get(current_day.isoformat(), 0)))
        current_day += timedelta(days=1)
    return series


def _hojas_filtradas_context(request: HttpRequest, *, usar_fechas: bool = True) -> dict:
    estado = request.GET.get("estado", "").strip()
    q = request.GET.get("q", "").strip()
    today = timezone.localdate()
    desde = _parse_dashboard_date(request.GET.get("desde"), today - timedelta(days=29)) if usar_fechas else None
    hasta = _parse_dashboard_date(request.GET.get("hasta"), today) if usar_fechas else None
    desde_dt = None
    hasta_dt = None
    if usar_fechas and desde and hasta:
        if desde > hasta:
            desde, hasta = hasta, desde
        desde_dt, _ = _date_bounds(desde)
        _, hasta_dt = _date_bounds(hasta)

    hojas_qs = HojaRuta.objects.annotate(
        remitos_total=Count("remitos", distinct=True),
        remitos_conformados=Count(
            "remitos",
            filter=Q(remitos__estado__in=(Remito.Estado.VALIDADO, Remito.Estado.OBSERVADO)),
            distinct=True,
        ),
        remitos_pendientes=Count(
            "remitos",
            filter=~Q(remitos__estado__in=(Remito.Estado.VALIDADO, Remito.Estado.OBSERVADO)),
            distinct=True,
        ),
    ).order_by("-created_at")
    if desde_dt and hasta_dt:
        hojas_qs = hojas_qs.filter(created_at__range=(desde_dt, hasta_dt))
    if estado:
        hojas_qs = hojas_qs.filter(estado=estado)
    if q:
        hojas_qs = hojas_qs.filter(nro_entrega__icontains=q)

    return {
        "estado": estado,
        "q": q,
        "desde": desde,
        "hasta": hasta,
        "desde_dt": desde_dt,
        "hasta_dt": hasta_dt,
        "hojas_qs": hojas_qs,
        "hojas": hojas_qs[:50],
    }


def _save_import_preview_file(request: HttpRequest, uploaded_file, kind: str) -> str:
    token = uuid.uuid4().hex
    original_name = Path(uploaded_file.name or "archivo").name
    if hasattr(uploaded_file, "seek"):
        try:
            uploaded_file.seek(0)
        except Exception:
            pass
    path = default_storage.save(f"import-preview/{request.user.pk}/{token}/{original_name}", uploaded_file)
    previews = request.session.get(IMPORT_PREVIEW_SESSION_KEY, {})
    previews[token] = {
        "path": path,
        "name": original_name,
        "kind": kind,
        "content_type": getattr(uploaded_file, "content_type", "") or "",
    }
    request.session[IMPORT_PREVIEW_SESSION_KEY] = previews
    request.session.modified = True
    return token


def _get_import_preview_meta(request: HttpRequest, token: str, kind: str) -> dict[str, str]:
    previews = request.session.get(IMPORT_PREVIEW_SESSION_KEY, {})
    meta = previews.get(token)
    if not meta or meta.get("kind") != kind:
        raise ValueError("No se encontro el archivo previsualizado. Seleccionalo nuevamente.")
    if not default_storage.exists(meta["path"]):
        raise ValueError("El archivo previsualizado ya no esta disponible. Seleccionalo nuevamente.")
    return meta


def _open_import_preview_file(request: HttpRequest, token: str, kind: str) -> File:
    meta = _get_import_preview_meta(request, token, kind)
    return File(default_storage.open(meta["path"], "rb"), name=meta["name"])


def _delete_import_preview_file(request: HttpRequest, token: str) -> None:
    previews = request.session.get(IMPORT_PREVIEW_SESSION_KEY, {})
    meta = previews.pop(token, None)
    if meta and default_storage.exists(meta["path"]):
        default_storage.delete(meta["path"])
    request.session[IMPORT_PREVIEW_SESSION_KEY] = previews
    request.session.modified = True


def _format_manual_remito(value: str) -> str:
    raw = (value or "").strip()
    if REMITO_MANUAL_PATTERN.match(raw):
        return raw

    if REMITO_MANUAL_DIGITS_PATTERN.match(raw):
        return f"{raw[:5]}-{raw[5:]}"

    digits = re.sub(r"\D+", "", raw)
    if digits:
        missing = 13 - len(digits)
        if missing > 0:
            raise ValueError(f"Al numero de remito le faltan {missing} digito(s). Formato esperado: 00009-00022221.")
        if len(digits) > 13:
            raise ValueError("El numero de remito tiene digitos de mas. Formato esperado: 00009-00022221.")

    raise ValueError("Formato de remito invalido. Usa 5 digitos, guion y 8 digitos. Ejemplo: 00009-00022221.")


def _extract_remito_oid_from_qr(value: str) -> str:
    raw = (value or "").strip()
    match = OID_PATTERN.search(raw)
    return match.group(0) if match else raw


def _find_remito_in_hoja(*, hoja: HojaRuta, remito_input: str, origen: str = "manual") -> Remito:
    raw = (remito_input or "").strip()
    if not raw:
        raise ValueError("Debes informar un remito para continuar.")

    if origen == "qr":
        remito_oid = _extract_remito_oid_from_qr(raw)
        normalized_oid = _normalize_code(remito_oid)
        for remito in hoja.remitos.only("id", "remito_uid"):
            if _normalize_code(remito.remito_uid) == normalized_oid:
                return remito
        raise ValueError("El QR escaneado no corresponde a un remito de esta hoja de ruta.")

    remito_numero = _format_manual_remito(raw)
    exact = hoja.remitos.filter(numero=remito_numero).first()
    if exact:
        return exact

    raise ValueError("No se encontro ese numero de remito en esta hoja de ruta.")


def _build_evidencia_file_context(evidencia: Evidencia) -> dict[str, str | bool]:
    archivo = evidencia.archivo
    if not archivo:
        return {"available": False, "url": "", "name": "", "is_image": False, "is_pdf": False}

    name = Path(archivo.name).name
    extension = Path(archivo.name).suffix.lower()
    try:
        url = archivo.url
    except ValueError:
        return {"available": False, "url": "", "name": name, "is_image": False, "is_pdf": False}

    return {
        "available": True,
        "url": url,
        "name": name,
        "is_image": extension in IMAGE_EXTENSIONS,
        "is_pdf": extension in PDF_EXTENSIONS,
    }


def _public_probe_rate_limit_key(*, request: HttpRequest, canal: str) -> str:
    ip = _normalize_code(_get_client_ip(request))
    return f"probe:{ip}:{canal}"


def _timeline_badge_class(tipo: str) -> str:
    return {
        EventoTrazabilidad.Tipo.IMPORTACION: "bg-secondary",
        EventoTrazabilidad.Tipo.APERTURA_LINK: "bg-info text-dark",
        EventoTrazabilidad.Tipo.CARGA_EVIDENCIA: "bg-success",
        EventoTrazabilidad.Tipo.INTENTO_FALLIDO: "bg-warning text-dark",
        EventoTrazabilidad.Tipo.VALIDACION: "bg-primary",
        EventoTrazabilidad.Tipo.RECHAZO: "bg-danger",
        EventoTrazabilidad.Tipo.CIERRE: "bg-dark",
    }.get(tipo, "bg-secondary")


def root_redirect(request: HttpRequest) -> HttpResponse:
    return redirect("dashboard")


def login_view(request: HttpRequest) -> HttpResponse:
    if request.user.is_authenticated:
        return redirect("dashboard")

    if request.method == "POST":
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            login(request, form.get_user())
            return redirect("dashboard")
    else:
        form = LoginForm(request)
    return render(request, "registration/login.html", {"form": form})


def logout_view(request: HttpRequest) -> HttpResponse:
    logout(request)
    return redirect("login")


@login_required
@user_passes_test(can_manage_users)
def panel_crear_usuario(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        form = UserCreateForm(request.POST)
        if form.is_valid():
            try:
                create_user_with_profile(
                    username=form.cleaned_data["username"],
                    email=form.cleaned_data.get("email", ""),
                    password=form.cleaned_data["password1"],
                    rol=form.cleaned_data["rol"],
                    share_logistica=bool(form.cleaned_data.get("share_logistica")),
                    share_cliente=bool(form.cleaned_data.get("share_cliente")),
                )
            except Exception as exc:
                form.add_error(None, str(exc))
            else:
                messages.success(request, "Usuario creado correctamente.")
                return redirect("panel-home")
    else:
        form = UserCreateForm()

    return render(request, "tracking/panel_crear_usuario.html", {"form": form})


@login_required
@user_passes_test(can_manage_users)
def panel_usuarios(request: HttpRequest) -> HttpResponse:
    users = User.objects.order_by("username")
    rows = [{"user": user, "profile": get_or_create_profile(user)} for user in users]
    return render(
        request,
        "tracking/panel_usuarios.html",
        {
            "rows": rows,
            "can_grant_staff": can_grant_staff(request.user),
        },
    )


@login_required
@user_passes_test(can_manage_users)
def panel_permisos(request: HttpRequest) -> HttpResponse:
    permisos = [
        {
            "accion": "Importar PDF",
            "deposito": "Si",
            "ventas": "No",
            "jefe": "Si",
            "otro": "No",
        },
        {
            "accion": "Revisar y validar evidencias",
            "deposito": "No",
            "ventas": "No",
            "jefe": "Si",
            "otro": "No",
        },
        {
            "accion": "Cerrar hoja",
            "deposito": "No",
            "ventas": "No",
            "jefe": "Si",
            "otro": "No",
        },
        {
            "accion": "Gestionar usuarios",
            "deposito": "No",
            "ventas": "No",
            "jefe": "Si",
            "otro": "No",
        },
        {
            "accion": "Compartir link logistica",
            "deposito": "Si",
            "ventas": "No",
            "jefe": "Si",
            "otro": "No",
        },
        {
            "accion": "Compartir link cliente",
            "deposito": "No",
            "ventas": "Si",
            "jefe": "Si",
            "otro": "No",
        },
    ]
    return render(request, "tracking/panel_permisos.html", {"permisos": permisos})


@login_required
@user_passes_test(can_manage_users)
def panel_editar_usuario(request: HttpRequest, user_id: int) -> HttpResponse:
    usuario = get_object_or_404(User, pk=user_id)
    profile = get_or_create_profile(usuario)
    actor_can_grant_staff = can_grant_staff(request.user)

    if request.method == "POST":
        form = UserUpdateForm(request.POST)
        if not actor_can_grant_staff:
            form.fields.pop("is_staff", None)
        if form.is_valid():
            is_staff_target = bool(form.cleaned_data.get("is_staff")) if actor_can_grant_staff else usuario.is_staff
            try:
                update_user_with_profile(
                    user=usuario,
                    username=form.cleaned_data["username"],
                    email=form.cleaned_data.get("email", ""),
                    rol=form.cleaned_data["rol"],
                    share_logistica=bool(form.cleaned_data.get("share_logistica")),
                    share_cliente=bool(form.cleaned_data.get("share_cliente")),
                    is_active=bool(form.cleaned_data.get("is_active")),
                    is_staff=is_staff_target,
                    password=form.cleaned_data.get("password", ""),
                )
            except Exception as exc:
                form.add_error(None, str(exc))
            else:
                messages.success(request, "Usuario actualizado correctamente.")
                return redirect("panel-usuarios")
    else:
        form = UserUpdateForm(
            initial={
                "username": usuario.username,
                "email": usuario.email,
                "rol": profile.rol,
                "share_logistica": profile.share_logistica,
                "share_cliente": profile.share_cliente,
                "is_active": usuario.is_active,
                "is_staff": usuario.is_staff,
            }
        )
        if not actor_can_grant_staff:
            form.fields.pop("is_staff", None)

    return render(
        request,
        "tracking/panel_editar_usuario.html",
        {
            "form": form,
            "usuario": usuario,
            "actor_can_grant_staff": actor_can_grant_staff,
        },
    )


@login_required
@user_passes_test(can_manage_users)
def panel_eliminar_usuario(request: HttpRequest, user_id: int) -> HttpResponse:
    usuario = get_object_or_404(User, pk=user_id)
    if usuario == request.user:
        messages.error(request, "No podes eliminar tu propio usuario.")
        return redirect("panel-usuarios")

    if request.method == "POST":
        form = UserDeleteForm(request.POST)
        if form.is_valid():
            delete_user_and_profile(user=usuario)
            messages.success(request, "Usuario eliminado correctamente.")
            return redirect("panel-usuarios")
    else:
        form = UserDeleteForm()

    return render(request, "tracking/panel_eliminar_usuario.html", {"form": form, "usuario": usuario})


@login_required
def panel_home(request: HttpRequest) -> HttpResponse:
    filtros = _hojas_filtradas_context(request, usar_fechas=False)
    paginator = Paginator(filtros["hojas_qs"], 20)
    page_obj = paginator.get_page(request.GET.get("page"))
    return render(
        request,
        "tracking/panel_home.html",
        {
            "hojas": page_obj.object_list,
            "page_obj": page_obj,
            "total_hojas": paginator.count,
            "estado": filtros["estado"],
            "q": filtros["q"],
            "estados": HojaRuta.Estado.choices,
            "can_import_pdf": can_import_pdf(request.user),
            "can_export_excel": can_import_pdf(request.user),
            "can_review_evidence": can_review_evidence(request.user),
            "can_manage_users": can_manage_users(request.user),
            "can_close_hoja": can_close_hoja(request.user),
        },
    )


@login_required
def dashboard(request: HttpRequest) -> HttpResponse:
    filtros = _hojas_filtradas_context(request)
    estado = filtros["estado"]
    q = filtros["q"]
    desde = filtros["desde"]
    hasta = filtros["hasta"]
    desde_dt = filtros["desde_dt"]
    hasta_dt = filtros["hasta_dt"]
    period_labels = []
    cursor = desde
    while cursor <= hasta:
        period_labels.append(cursor.isoformat())
        cursor += timedelta(days=1)

    accesos_inexistentes = IntentoAccesoPortal.objects.filter(
        fecha_evento__range=(desde_dt, hasta_dt),
        motivo=IntentoAccesoPortal.Motivo.HOJA_INEXISTENTE,
    )

    hojas_grafico_qs = HojaRuta.objects.filter(created_at__range=(desde_dt, hasta_dt))
    if estado:
        hojas_grafico_qs = hojas_grafico_qs.filter(estado=estado)
    if q:
        hojas_grafico_qs = hojas_grafico_qs.filter(nro_entrega__icontains=q)
    hojas_por_dia = {
        row["day"].isoformat(): row["total"]
        for row in hojas_grafico_qs.annotate(day=TruncDate("created_at")).values("day").annotate(total=Count("id"))
    }

    hojas_metricas = filtros["hojas_qs"]
    remitos_metricas = Remito.objects.filter(hoja_ruta__in=hojas_metricas)
    evidencias_metricas = Evidencia.objects.filter(hoja_ruta__in=hojas_metricas)
    hojas_cargadas_periodo = hojas_metricas.count()
    remitos_total_periodo = remitos_metricas.count()
    remitos_con_evidencia_periodo = evidencias_metricas.values("remito_id").distinct().count()
    remitos_evidencia_aprobada_periodo = evidencias_metricas.filter(
        estado_validacion=Evidencia.EstadoValidacion.VALIDADA
    ).values("remito_id").distinct().count()
    evidencias_pendientes_auditoria = evidencias_metricas.filter(
        estado_validacion=Evidencia.EstadoValidacion.PENDIENTE
    ).count()
    hr_no_cargadas = accesos_inexistentes.count()
    hr_no_cargadas_unicas = accesos_inexistentes.values("oid").distinct().count()
    return render(
        request,
        "tracking/dashboard.html",
        {
            "estado": estado,
            "q": q,
            "desde": desde.isoformat(),
            "hasta": hasta.isoformat(),
            "estados": HojaRuta.Estado.choices,
            "dashboard": {
                "labels": period_labels,
                "hojas_cargadas": _build_day_series(desde, hasta, hojas_por_dia),
            },
            "kpis": {
                "hojas_cargadas_periodo": hojas_cargadas_periodo,
                "remitos_total_periodo": remitos_total_periodo,
                "remitos_con_evidencia_periodo": remitos_con_evidencia_periodo,
                "remitos_evidencia_aprobada_periodo": remitos_evidencia_aprobada_periodo,
                "evidencias_pendientes_auditoria": evidencias_pendientes_auditoria,
                "hr_no_cargadas": hr_no_cargadas,
                "hr_no_cargadas_unicas": hr_no_cargadas_unicas,
            },
        },
    )


@login_required
def panel_auditoria_hr_no_cargadas(request: HttpRequest) -> HttpResponse:
    if not can_review_evidence(request.user):
        messages.error(request, "No tenes permisos para revisar auditoria de HR no cargadas.")
        return redirect("panel-home")

    q = request.GET.get("q", "").strip()
    canal = request.GET.get("canal", "").strip()
    today = timezone.localdate()
    desde = _parse_dashboard_date(request.GET.get("desde"), today - timedelta(days=29))
    hasta = _parse_dashboard_date(request.GET.get("hasta"), today)
    if desde > hasta:
        desde, hasta = hasta, desde

    desde_dt, _ = _date_bounds(desde)
    _, hasta_dt = _date_bounds(hasta)
    intentos_qs = IntentoAccesoPortal.objects.filter(
        fecha_evento__range=(desde_dt, hasta_dt),
        motivo=IntentoAccesoPortal.Motivo.HOJA_INEXISTENTE,
    ).order_by("-fecha_evento")
    if q:
        intentos_qs = intentos_qs.filter(oid__icontains=q)
    if canal:
        intentos_qs = intentos_qs.filter(canal=canal)

    canales = (
        IntentoAccesoPortal.objects.filter(motivo=IntentoAccesoPortal.Motivo.HOJA_INEXISTENTE)
        .exclude(canal="")
        .values_list("canal", flat=True)
        .distinct()
        .order_by("canal")
    )
    intentos = intentos_qs[:200]
    return render(
        request,
        "tracking/panel_auditoria_hr_no_cargadas.html",
        {
            "intentos": intentos,
            "q": q,
            "canal": canal,
            "canales": canales,
            "desde": desde.isoformat(),
            "hasta": hasta.isoformat(),
            "total_intentos": intentos_qs.count(),
            "total_hr_unicas": intentos_qs.values("oid").distinct().count(),
        },
    )


@login_required
def panel_hoja_detalle(request: HttpRequest, oid: str) -> HttpResponse:
    hoja = get_object_or_404(HojaRuta, oid=oid)
    remito_estado = request.GET.get("remito_estado", "").strip()
    q = request.GET.get("q", "").strip()

    remitos_qs = hoja.remitos.order_by("id")
    if remito_estado:
        remitos_qs = remitos_qs.filter(estado=remito_estado)
    if q:
        remitos_qs = remitos_qs.filter(numero__icontains=q)

    remitos = remitos_qs[:200]
    evidencias = Evidencia.objects.filter(hoja_ruta=hoja).select_related("remito").order_by("-fecha_carga")[:30]
    intentos = hoja.intentos.select_related("remito").order_by("-fecha_evento")[:30]
    profile = get_or_create_profile(request.user)
    logistica_url = request.build_absolute_uri(f"/conformados/logistica/{hoja.oid}/")
    cliente_url = request.build_absolute_uri(f"/conformados/cliente/{hoja.oid}/")
    return render(
        request,
        "tracking/panel_hoja_detalle.html",
        {
            "hoja": hoja,
            "remitos": remitos,
            "evidencias": evidencias,
            "intentos": intentos,
            "remito_estado": remito_estado,
            "q": q,
            "remito_estados": Remito.Estado.choices,
            "profile": profile,
            "can_share_logistica": profile.can_share_logistica(),
            "can_share_cliente": profile.can_share_cliente(),
            "logistica_url": logistica_url,
            "cliente_url": cliente_url,
        },
    )


@login_required
def panel_evidencias(request: HttpRequest) -> HttpResponse:
    if not can_review_evidence(request.user):
        messages.error(request, "No tenes permisos para revisar evidencias.")
        return redirect("panel-home")

    remito_q = request.GET.get("remito", "").strip()
    estado = request.GET.get("estado", "").strip()

    evidencias_qs = Evidencia.objects.select_related("hoja_ruta", "remito").order_by("-fecha_carga")
    if remito_q:
        evidencias_qs = evidencias_qs.filter(remito__numero__icontains=remito_q)
    if estado:
        evidencias_qs = evidencias_qs.filter(estado_validacion=estado)

    evidencias = evidencias_qs[:200]
    return render(
        request,
        "tracking/panel_evidencias.html",
        {
            "evidencias": evidencias,
            "remito": remito_q,
            "estado": estado,
            "estados": Evidencia.EstadoValidacion.choices,
        },
    )


@login_required
def panel_auditoria_remitos(request: HttpRequest) -> HttpResponse:
    if not can_review_evidence(request.user):
        messages.error(request, "No tenes permisos para revisar auditoría de remitos.")
        return redirect("panel-home")

    q = request.GET.get("q", "").strip()
    estado = request.GET.get("estado", "").strip()
    conformado = request.GET.get("conformado", "").strip()
    hoja = request.GET.get("hoja", "").strip()

    remitos_qs = (
        Remito.objects.select_related("hoja_ruta")
        .annotate(
            evidencias_total=Count("evidencias", distinct=True),
            intentos_total=Count("intentos", distinct=True),
            fecha_ultima_evidencia=Max("evidencias__fecha_carga"),
            fecha_ultimo_evento=Max("eventos__fecha_evento"),
        )
        .order_by("-hoja_ruta__fecha", "-created_at", "numero")
    )

    if q:
        remitos_qs = remitos_qs.filter(
            Q(numero__icontains=q)
            | Q(cliente__icontains=q)
            | Q(remito_uid__icontains=q)
            | Q(hoja_ruta__nro_entrega__icontains=q)
        )
    if estado:
        remitos_qs = remitos_qs.filter(estado=estado)
    if hoja:
        remitos_qs = remitos_qs.filter(hoja_ruta__nro_entrega__icontains=hoja)
    if conformado == "si":
        remitos_qs = remitos_qs.filter(evidencias_total__gt=0)
    elif conformado == "no":
        remitos_qs = remitos_qs.filter(evidencias_total=0)

    return render(
        request,
        "tracking/panel_auditoria_remitos.html",
        {
            "remitos": remitos_qs,
            "q": q,
            "estado": estado,
            "conformado": conformado,
            "hoja": hoja,
            "estados": Remito.Estado.choices,
            "conformados": (("", "Todos"), ("si", "Con conformado"), ("no", "Sin conformado")),
        },
    )


@login_required
def panel_auditoria_remito_detalle(request: HttpRequest, remito_id: int) -> HttpResponse:
    if not can_review_evidence(request.user):
        messages.error(request, "No tenes permisos para revisar auditoría de remitos.")
        return redirect("panel-home")

    remito = get_object_or_404(Remito.objects.select_related("hoja_ruta"), pk=remito_id)
    evidencias = list(remito.evidencias.select_related("hoja_ruta", "remito").order_by("-fecha_carga"))
    intentos = list(remito.intentos.select_related("hoja_ruta", "remito").order_by("-fecha_evento"))
    timeline_qs = (
        remito.hoja_ruta.eventos.select_related("remito")
        .filter(Q(remito=remito) | Q(remito__isnull=True))
        .order_by("fecha_evento")
    )

    # Asociar evidencias e intentos a cada evento por proximidad temporal
    timeline = []
    for evento in timeline_qs:
        # buscar evidencias cercanas al evento (2 minutos) y construir contexto de archivo
        raw_matched_evidencias = [e for e in evidencias if abs((e.fecha_carga - evento.fecha_evento).total_seconds()) <= 120]
        matched_evidencias = [
            {"evidencia": e, "archivo": _build_evidencia_file_context(e)} for e in raw_matched_evidencias
        ]
        # buscar intentos cercanos al evento (2 minutos)
        matched_intentos = [i for i in intentos if abs((i.fecha_evento - evento.fecha_evento).total_seconds()) <= 120]

        timeline.append(
            {
                "evento": evento,
                "badge_class": _timeline_badge_class(evento.tipo),
                "evidencias": matched_evidencias,
                "intentos": matched_intentos,
            }
        )

    evidencia_rows = [
        {
            "evidencia": evidencia,
            "archivo": _build_evidencia_file_context(evidencia),
        }
        for evidencia in evidencias
    ]

    return render(
        request,
        "tracking/panel_auditoria_remito_detalle.html",
        {
            "remito": remito,
            "evidencias": evidencia_rows,
            "intentos": intentos,
            "timeline": timeline,
            "fecha_ultima_evidencia": evidencias[0].fecha_carga if evidencias else None,
            "total_evidencias": len(evidencias),
            "total_intentos": len(intentos),
        },
    )


@login_required
def panel_importar_pdf(request: HttpRequest) -> HttpResponse:
    if not can_import_pdf(request.user):
        messages.error(request, "No tenes permisos para importar PDF.")
        return redirect("panel-home")

    preview = None
    preview_token = ""
    preview_file_name = ""
    if request.method == "POST":
        action = request.POST.get("action", "preview")
        token = request.POST.get("preview_token", "").strip()

        if action == "import" and token and "pdf_file" not in request.FILES:
            form = ImportPdfForm()
            try:
                with _open_import_preview_file(request, token, "pdf") as pdf_file:
                    hoja = import_hoja_ruta_pdf(pdf_file)
            except Exception as exc:
                form.add_error("pdf_file", str(exc))
                preview_token = token
                try:
                    preview_file_name = _get_import_preview_meta(request, token, "pdf")["name"]
                except Exception:
                    preview_file_name = ""
            else:
                _delete_import_preview_file(request, token)
                messages.success(request, f"Hoja {hoja.nro_entrega} importada correctamente.")
                return redirect("panel-home")
        else:
            form = ImportPdfForm(request.POST, request.FILES)
            if form.is_valid():
                pdf_file = form.cleaned_data["pdf_file"]
                try:
                    text = extract_text_from_pdf(pdf_file)
                    oid = extract_oid_from_qr(pdf_file)
                    parsed = parse_hoja_ruta_pdf(text, oid=oid)
                    _validate_parsed_hoja(parsed)
                except Exception as exc:
                    form.add_error("pdf_file", str(exc))
                else:
                    if action == "import":
                        hoja = import_hoja_ruta_pdf(pdf_file)
                        messages.success(request, f"Hoja {hoja.nro_entrega} importada correctamente.")
                        return redirect("panel-home")

                    preview = parsed
                    preview_token = _save_import_preview_file(request, pdf_file, "pdf")
                    preview_file_name = Path(pdf_file.name or "").name
    else:
        form = ImportPdfForm()

    return render(
        request,
        "tracking/importar_pdf.html",
        {"form": form, "preview": preview, "preview_token": preview_token, "preview_file_name": preview_file_name},
    )


@login_required
def panel_importar_excel(request: HttpRequest) -> HttpResponse:
    if not can_import_pdf(request.user):
        messages.error(request, "No tenes permisos para importar archivos.")
        return redirect("panel-home")

    preview = None
    preview_token = ""
    preview_file_name = ""
    if request.method == "POST":
        action = request.POST.get("action", "preview")
        token = request.POST.get("preview_token", "").strip()

        if action == "import" and token and "archivo" not in request.FILES:
            form = ImportSpreadsheetForm()
            try:
                with _open_import_preview_file(request, token, "tabular") as archivo:
                    hojas, remitos = import_tabular_file(archivo, archivo)
            except Exception as exc:
                form.add_error("archivo", str(exc))
                preview_token = token
                try:
                    preview_file_name = _get_import_preview_meta(request, token, "tabular")["name"]
                except Exception:
                    preview_file_name = ""
            else:
                _delete_import_preview_file(request, token)
                messages.success(request, f"Importacion completada: {hojas} hoja(s), {remitos} remito(s).")
                return redirect("panel-home")
        else:
            form = ImportSpreadsheetForm(request.POST, request.FILES)
            if form.is_valid():
                archivo = form.cleaned_data["archivo"]
                try:
                    parsed = parse_tabular_file(archivo)
                    _validate_parsed_hoja(parsed)
                except Exception as exc:
                    form.add_error("archivo", str(exc))
                else:
                    if action == "import":
                        hojas, remitos = import_tabular_file(archivo, archivo)
                        messages.success(request, f"Importacion completada: {hojas} hoja(s), {remitos} remito(s).")
                        return redirect("panel-home")

                    preview = parsed
                    preview_token = _save_import_preview_file(request, archivo, "tabular")
                    preview_file_name = Path(archivo.name or "").name
    else:
        form = ImportSpreadsheetForm()

    return render(
        request,
        "tracking/importar_excel.html",
        {"form": form, "preview": preview, "preview_token": preview_token, "preview_file_name": preview_file_name},
    )


@login_required
def _panel_exportar_excel_detallado_legacy(request: HttpRequest) -> HttpResponse:
    if not can_import_pdf(request.user):
        messages.error(request, "No tenes permisos para exportar archivos.")
        return redirect("panel-home")

    # Crear workbook
    wb = Workbook()
    ws_default = wb.active
    ws_default.title = "Info"

    # Obtener todas las hojas de ruta
    hojas = HojaRuta.objects.all().prefetch_related("remitos", "evidencias").order_by("-created_at")

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    if not hojas.exists():
        # Si no hay hojas, mostrar mensaje en la hoja por defecto
        ws_default[f"A1"] = "No hay hojas de ruta cargadas"
        ws_default[f"A1"].font = Font(bold=True, size=12)
        ws_default.column_dimensions["A"].width = 30
    else:
        # Limpiar la hoja por defecto si hay datos
        wb.remove(ws_default)
        
        for hoja in hojas:
            # Crear hoja por cada HojaRuta
            ws = wb.create_sheet(title=f"{hoja.nro_entrega[:31]}")

            # Encabezado con datos de la hoja de ruta
            row = 1
            ws[f"A{row}"] = "HOJA DE RUTA"
            ws[f"A{row}"].font = Font(bold=True, size=12)
            ws.merge_cells(f"A{row}:C{row}")

            row += 1
            ws[f"A{row}"] = "OID:"
            ws[f"B{row}"] = str(hoja.oid)
            row += 1
            ws[f"A{row}"] = "Nro Entrega:"
            ws[f"B{row}"] = hoja.nro_entrega
            row += 1
            ws[f"A{row}"] = "Fecha:"
            ws[f"B{row}"] = hoja.fecha.strftime("%d/%m/%Y") if hoja.fecha else "-"
            row += 1
            ws[f"A{row}"] = "Estado:"
            ws[f"B{row}"] = hoja.get_estado_display()
            row += 1
            ws[f"A{row}"] = "Transporte Tipo:"
            ws[f"B{row}"] = hoja.transporte_tipo or "-"
            row += 1
            ws[f"A{row}"] = "Flete:"
            ws[f"B{row}"] = hoja.flete or "-"
            row += 1
            ws[f"A{row}"] = "Chofer:"
            ws[f"B{row}"] = hoja.chofer or "-"
            row += 1
            ws[f"A{row}"] = "Acompañante:"
            ws[f"B{row}"] = hoja.acompanante or "-"
            row += 1
            ws[f"A{row}"] = "Transporte:"
            ws[f"B{row}"] = hoja.transporte or "-"
            row += 1
            ws[f"A{row}"] = "Creado:"
            ws[f"B{row}"] = hoja.created_at.strftime("%d/%m/%Y %H:%M") if hoja.created_at else "-"

            # Tabla de remitos
            row += 2
            headers = ["Remito", "OID Remito", "Cliente", "Subcliente", "Dirección", "Observación", "Estado", "Evidencias"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment

            row += 1
            for remito in hoja.remitos.all():
                evidencias_count = remito.evidencias.count()
                ws.cell(row=row, column=1, value=remito.numero)
                ws.cell(row=row, column=2, value=remito.remito_uid)
                ws.cell(row=row, column=3, value=remito.cliente)
                ws.cell(row=row, column=4, value=remito.subcliente or "-")
                ws.cell(row=row, column=5, value=remito.direccion)
                ws.cell(row=row, column=6, value=remito.observacion or "-")
                ws.cell(row=row, column=7, value=remito.get_estado_display())
                ws.cell(row=row, column=8, value=evidencias_count)

                # Aplicar estilos a las celdas
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).alignment = left_alignment

                row += 1

            # Ajustar ancho de columnas
            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 25
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 20
            ws.column_dimensions["E"].width = 25
            ws.column_dimensions["F"].width = 20
            ws.column_dimensions["G"].width = 18
            ws.column_dimensions["H"].width = 12

    # Generar archivo en memoria
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Retornar archivo para descargar
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="hojas-ruta.xlsx"'
    return response


@login_required
def panel_exportar_excel(request: HttpRequest) -> HttpResponse:
    if not can_import_pdf(request.user):
        messages.error(request, "No tenes permisos para exportar archivos.")
        return redirect("panel-home")

    estado = request.GET.get("estado", "").strip()
    q = request.GET.get("q", "").strip()
    desde = parse_date(request.GET.get("desde", ""))
    hasta = parse_date(request.GET.get("hasta", ""))
    if desde and hasta and desde > hasta:
        desde, hasta = hasta, desde

    hojas = HojaRuta.objects.annotate(
        remitos_total=Count("remitos", distinct=True),
        remitos_con_evidencia=Count("remitos__evidencias__remito", distinct=True),
        remitos_evidencia_aprobada=Count(
            "remitos__evidencias__remito",
            filter=Q(remitos__evidencias__estado_validacion=Evidencia.EstadoValidacion.VALIDADA),
            distinct=True,
        ),
        evidencias_sin_auditar=Count(
            "evidencias",
            filter=Q(evidencias__estado_validacion=Evidencia.EstadoValidacion.PENDIENTE),
            distinct=True,
        ),
        evidencias_rechazadas=Count(
            "evidencias",
            filter=Q(evidencias__estado_validacion=Evidencia.EstadoValidacion.RECHAZADA),
            distinct=True,
        ),
    ).order_by("-created_at")
    if desde:
        desde_dt, _ = _date_bounds(desde)
        hojas = hojas.filter(created_at__gte=desde_dt)
    if hasta:
        _, hasta_dt = _date_bounds(hasta)
        hojas = hojas.filter(created_at__lte=hasta_dt)
    if estado:
        hojas = hojas.filter(estado=estado)
    if q:
        hojas = hojas.filter(nro_entrega__icontains=q)

    wb = Workbook()
    ws = wb.active
    ws.title = "Hojas de ruta"
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws["A1"] = "Listado de hojas de ruta filtradas"
    ws["A1"].font = Font(bold=True, size=13)
    desde_label = desde.strftime("%d/%m/%Y") if desde else "sin inicio"
    hasta_label = hasta.strftime("%d/%m/%Y") if hasta else "sin fin"
    ws["A2"] = f"Periodo de importacion: {desde_label} a {hasta_label}"
    ws["A3"] = f"Estado: {dict(HojaRuta.Estado.choices).get(estado, 'Todos')}"
    ws["A4"] = f"Busqueda entrega: {q or '-'}"
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")
    ws.merge_cells("A3:D3")
    ws.merge_cells("A4:D4")

    headers = [
        "Nro entrega",
        "OID hoja",
        "Fecha hoja",
        "Estado",
        "Remitos totales",
        "Remitos con evidencia",
        "Remitos con evidencia aprobada",
        "Evidencias sin auditar",
        "Evidencias rechazadas",
        "Transporte tipo",
        "Flete",
        "Chofer",
        "Acompanante",
        "Transporte",
        "Fecha importacion",
    ]
    header_row = 6
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    row = header_row + 1
    for hoja in hojas:
        values = [
            hoja.nro_entrega,
            str(hoja.oid),
            hoja.fecha.strftime("%d/%m/%Y") if hoja.fecha else "",
            hoja.get_estado_display(),
            hoja.remitos_total,
            hoja.remitos_con_evidencia,
            hoja.remitos_evidencia_aprobada,
            hoja.evidencias_sin_auditar,
            hoja.evidencias_rechazadas,
            hoja.transporte_tipo,
            hoja.flete,
            hoja.chofer,
            hoja.acompanante,
            hoja.transporte,
            timezone.localtime(hoja.created_at).strftime("%d/%m/%Y %H:%M") if hoja.created_at else "",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = left_alignment if col in {1, 2, 4, 10, 11, 12, 13, 14} else center_alignment
        row += 1

    if row == header_row + 1:
        ws.cell(row=row, column=1, value="No hay hojas de ruta para los filtros seleccionados.")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))

    widths = [18, 40, 14, 14, 16, 20, 28, 20, 20, 20, 18, 22, 22, 22, 20]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=header_row, column=index).column_letter].width = width
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{ws.cell(row=max(row - 1, header_row), column=len(headers)).coordinate}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="listado-hojas-ruta.xlsx"'
    return response


def conformados_portal(request: HttpRequest, canal: str, oid: str) -> HttpResponse:
    oid_text = str(oid).strip()

    if not OID_PATTERN.fullmatch(oid_text):
        try:
            _check_rate_limit(
                key=_public_probe_rate_limit_key(request=request, canal=canal),
                limit=settings.PUBLIC_LINK_INVALID_RATE_LIMIT_COUNT,
                window_seconds=settings.PUBLIC_LINK_INVALID_RATE_LIMIT_WINDOW_SECONDS,
            )
        except ValueError:
            return HttpResponse("Demasiados intentos. Intenta mas tarde.", status=429)

        registrar_intento_acceso_portal(
            canal=canal,
            oid=oid_text,
            motivo="oid_invalido",
            ip_address=_get_client_ip(request),
            user_agent=request.META.get("HTTP_USER_AGENT", "")[:500],
            path=request.get_full_path()[:255],
            detalle="El valor solicitado no tiene formato de oid valido.",
        )
        return render(request, "tracking/estado_hoja.html", {"estado": "inexistente", "canal": canal, "oid": oid_text})

    hoja = HojaRuta.objects.filter(oid=oid_text).first()
    if not hoja:
        intento = registrar_intento_acceso_portal(
            canal=canal,
            oid=oid_text,
            motivo="hoja_inexistente",
            ip_address=_get_client_ip(request),
            user_agent=request.META.get("HTTP_USER_AGENT", "")[:500],
            path=request.get_full_path()[:255],
            detalle="Se intento abrir una hoja de ruta que no existe en el sistema.",
        )
        try:
            send_public_access_alert(intento=intento)
        except Exception:
            pass
        return render(request, "tracking/estado_hoja.html", {"estado": "inexistente", "canal": canal, "oid": oid_text})
    if hoja.estado == HojaRuta.Estado.CERRADA:
        return render(request, "tracking/estado_hoja.html", {"estado": "cerrada", "canal": canal, "oid": oid_text})

    remito_query = request.GET.get("remito", "").strip()
    remito_origen = request.GET.get("origen", "manual").strip()
    if remito_origen not in {"manual", "qr"}:
        remito_origen = "manual"
    modo = request.GET.get("modo", "evidencia").strip()
    if modo not in {"evidencia", "no_entregado"}:
        modo = "evidencia"

    remito_seleccionado = None
    remito_error = ""
    if remito_query:
        try:
            remito_seleccionado = _find_remito_in_hoja(hoja=hoja, remito_input=remito_query, origen=remito_origen)
            if remito_origen == "qr":
                remito_query = remito_seleccionado.numero
                remito_origen = "manual"
        except ValueError as exc:
            remito_error = str(exc)

    remitos = hoja.remitos.order_by("id")
    return render(
        request,
        "tracking/conformados_portal.html",
        {
            "hoja": hoja,
            "remitos": remitos,
            "evidencia_form": EvidenciaForm(),
            "no_entregado_form": NoEntregadoForm(),
            "canal": canal,
            "remito_query": remito_query,
            "remito_origen": remito_origen,
            "remito_seleccionado": remito_seleccionado,
            "remito_tiene_evidencia": bool(remito_seleccionado and remito_seleccionado.evidencias.exists()),
            "remito_error": remito_error,
            "modo": modo,
            "evidencia_limits": _evidencia_limits_context(),
        },
    )


def subir_evidencia(request: HttpRequest, canal: str, oid: str) -> HttpResponse:
    hoja = get_object_or_404(HojaRuta, oid=oid)
    if hoja.estado == HojaRuta.Estado.CERRADA:
        return render(request, "tracking/estado_hoja.html", {"estado": "cerrada", "canal": canal, "oid": oid})

    form = EvidenciaForm(request.POST, request.FILES)
    try:
        remito = _find_remito_in_hoja(hoja=hoja, remito_input=request.POST.get("remito_uid", ""), origen="qr")
    except ValueError as exc:
        messages.error(request, str(exc))
        remito_query = request.POST.get("remito_uid", "").strip()
        return redirect(f"/conformados/{canal}/{oid}/?remito={remito_query}&modo=evidencia")

    if form.is_valid():
        try:
            _check_rate_limit(
                key=_rate_limit_key(action="evidencia", request=request, canal=canal, oid=oid, remito_uid=remito.remito_uid),
                limit=settings.EVIDENCIA_RATE_LIMIT_COUNT,
                window_seconds=settings.EVIDENCIA_RATE_LIMIT_WINDOW_SECONDS,
            )
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect(f"/conformados/{canal}/{oid}/?remito={remito.numero}&modo=evidencia")

        try:
            registrar_evidencia(
                hoja=hoja,
                remito=remito,
                canal=canal,
                archivo=form.cleaned_data["archivo"],
                comentario=form.cleaned_data.get("comentario", ""),
                origen=form.cleaned_data.get("origen", ""),
                permitir_duplicada=bool(form.cleaned_data.get("confirmar_duplicada")),
            )
        except Exception as exc:
            form.add_error(None, str(exc))
        else:
            messages.success(request, "Evidencia cargada correctamente.")
            return redirect("conformados-portal", canal=canal, oid=oid)

    remitos = hoja.remitos.order_by("id")
    return render(
        request,
        "tracking/conformados_portal.html",
        {
            "hoja": hoja,
            "remitos": remitos,
            "evidencia_form": form,
            "no_entregado_form": NoEntregadoForm(),
            "canal": canal,
            "remito_query": remito.numero,
            "remito_origen": "manual",
            "remito_seleccionado": remito,
            "remito_tiene_evidencia": remito.evidencias.exists(),
            "remito_error": "",
            "modo": "evidencia",
            "evidencia_limits": _evidencia_limits_context(),
        },
    )


def no_entregado(request: HttpRequest, canal: str, oid: str) -> HttpResponse:
    hoja = get_object_or_404(HojaRuta, oid=oid)
    if hoja.estado == HojaRuta.Estado.CERRADA:
        return render(request, "tracking/estado_hoja.html", {"estado": "cerrada", "canal": canal, "oid": oid})

    form = NoEntregadoForm(request.POST)
    try:
        remito = _find_remito_in_hoja(hoja=hoja, remito_input=request.POST.get("remito_uid", ""), origen="qr")
    except ValueError as exc:
        messages.error(request, str(exc))
        remito_query = request.POST.get("remito_uid", "").strip()
        return redirect(f"/conformados/{canal}/{oid}/?remito={remito_query}&modo=no_entregado")

    if form.is_valid():
        try:
            _check_rate_limit(
                key=_rate_limit_key(action="no-entregado", request=request, canal=canal, oid=oid, remito_uid=remito.remito_uid),
                limit=settings.NO_ENTREGADO_RATE_LIMIT_COUNT,
                window_seconds=settings.NO_ENTREGADO_RATE_LIMIT_WINDOW_SECONDS,
            )
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect(f"/conformados/{canal}/{oid}/?remito={remito.numero}&modo=no_entregado")

        try:
            registrar_intento_no_entregado(
                hoja=hoja,
                remito=remito,
                canal=canal,
                motivo=form.cleaned_data["motivo"],
                comentario=form.cleaned_data.get("comentario", ""),
            )
        except Exception as exc:
            form.add_error(None, str(exc))
        else:
            messages.success(request, "Intento de no entrega registrado.")
            return redirect("conformados-portal", canal=canal, oid=oid)

    remitos = hoja.remitos.order_by("id")
    return render(
        request,
        "tracking/conformados_portal.html",
        {
            "hoja": hoja,
            "remitos": remitos,
            "evidencia_form": EvidenciaForm(),
            "no_entregado_form": form,
            "canal": canal,
            "remito_query": remito.numero,
            "remito_origen": "manual",
            "remito_seleccionado": remito,
            "remito_tiene_evidencia": remito.evidencias.exists(),
            "remito_error": "",
            "modo": "no_entregado",
            "evidencia_limits": _evidencia_limits_context(),
        },
    )


@login_required
def validar_evidencia(request: HttpRequest, evidencia_id: int) -> HttpResponse:
    if not can_review_evidence(request.user):
        messages.error(request, "No tenes permisos para validar evidencias.")
        return redirect("panel-home")

    evidencia = get_object_or_404(Evidencia.objects.select_related("hoja_ruta", "remito"), pk=evidencia_id)
    if request.method == "POST":
        form = ValidacionEvidenciaForm(request.POST)
        if form.is_valid():
            try:
                validar_evidencia_service(
                    evidencia=evidencia,
                    estado=form.cleaned_data["estado"],
                    comentario=form.cleaned_data.get("comentario", ""),
                )
            except Exception as exc:
                form.add_error(None, str(exc))
            else:
                messages.success(request, "Evidencia validada correctamente.")
                return redirect("panel-evidencias")
    else:
        form = ValidacionEvidenciaForm(initial={"estado": evidencia.estado_validacion})

    return render(
        request,
        "tracking/validar_evidencia.html",
        {"evidencia": evidencia, "form": form, "archivo": _build_evidencia_file_context(evidencia)},
    )


@login_required
def cerrar_hoja(request: HttpRequest, oid: str) -> HttpResponse:
    if not can_close_hoja(request.user):
        messages.error(request, "No tenes permisos para cerrar hojas.")
        return redirect("panel-home")

    hoja = get_object_or_404(HojaRuta, oid=oid)
    remitos_pendientes = hoja.remitos.exclude(
        estado__in=(Remito.Estado.VALIDADO, Remito.Estado.OBSERVADO)
    ).count()
    puede_cerrar = remitos_pendientes == 0
    if request.method == "POST":
        form = CierreHojaForm(request.POST)
        if form.is_valid():
            try:
                cerrar_hoja_ruta(hoja=hoja, comentario=form.cleaned_data.get("comentario", ""))
            except Exception as exc:
                form.add_error(None, str(exc))
            else:
                messages.success(request, "Hoja cerrada correctamente.")
                return redirect("panel-home")
    else:
        form = CierreHojaForm()

    return render(
        request,
        "tracking/cerrar_hoja.html",
        {"hoja": hoja, "form": form, "puede_cerrar": puede_cerrar, "remitos_pendientes": remitos_pendientes},
    )

from __future__ import annotations

import csv
import uuid
from collections import OrderedDict
from dataclasses import dataclass
from typing import Any, Iterable

from openpyxl import load_workbook

from tracking.services.import_pdf import RemitoData, _import_parsed_hoja, _normalize_value, _parse_date, _validate_parsed_hoja

TABULAR_REQUIRED_COLUMNS = {
    "oid",
    "nro_entrega",
    "fecha",
    "cliente",
    "remito",
    "remito_oid",
    "direccion",
}
TABULAR_OPTIONAL_COLUMNS = {
    "subcliente",
    "observacion",
    "transporte_tipo",
    "flete",
    "chofer",
    "acompanante",
    "transporte",
}


@dataclass
class TabularImportResult:
    hoja_count: int
    remito_count: int


def _normalize_header(value: str) -> str:
    return _normalize_value((value or "").replace("\ufeff", "")).lower().strip()


def _rows_to_dicts(headers: list[str], rows: Iterable[Iterable[Any]]) -> list[dict[str, str]]:
    result: list[dict[str, str]] = []
    for row in rows:
        values = list(row)
        mapped = {}
        for index, header in enumerate(headers):
            raw = values[index] if index < len(values) else ""
            mapped[header] = _normalize_value(str(raw)) if raw is not None else ""
        if any(value for value in mapped.values()):
            result.append(mapped)
    return result


def _group_rows_by_oid(rows: list[dict[str, str]]) -> OrderedDict[str, list[dict[str, str]]]:
    groups: OrderedDict[str, list[dict[str, str]]] = OrderedDict()
    for row in rows:
        oid = row.get("oid", "").strip()
        if not oid:
            raise ValueError("Cada fila debe incluir un oid.")
        groups.setdefault(oid, []).append(row)
    return groups


def _build_parsed_from_rows(rows: list[dict[str, str]]) -> dict[str, Any]:
    groups = _group_rows_by_oid(rows)
    if len(groups) != 1:
        raise ValueError("El archivo debe contener una sola hoja de ruta por importacion.")

    oid_text, grouped_rows = next(iter(groups.items()))
    oid = uuid.UUID(oid_text)
    first = grouped_rows[0]

    remitos: list[RemitoData] = []
    seen: set[str] = set()
    for row in grouped_rows:
        remito_numero = row.get("remito", "").strip()
        remito_oid = row.get("remito_oid", "").strip()
        if not remito_numero:
            raise ValueError("Cada fila debe incluir un numero de remito.")
        if not remito_oid:
            raise ValueError(f"Remito {remito_numero}: falta remito_oid.")
        try:
            remito_oid = str(uuid.UUID(remito_oid))
        except ValueError as exc:
            raise ValueError(f"Remito {remito_numero}: remito_oid debe ser un OID/UUID valido.") from exc
        if remito_oid in seen:
            raise ValueError(f"OID de remito duplicado detectado en archivo: {remito_oid}.")
        seen.add(remito_oid)

        cliente = row.get("cliente", "").strip()
        direccion = row.get("direccion", "").strip()
        if not cliente:
            raise ValueError(f"Remito {remito_numero}: falta cliente.")
        if not direccion:
            raise ValueError(f"Remito {remito_numero}: falta direccion.")

        remitos.append(
            RemitoData(
                remito_uid=remito_oid,
                numero=remito_numero,
                cliente=cliente,
                subcliente=row.get("subcliente", "").strip(),
                direccion=direccion,
                observacion=row.get("observacion", "").strip(),
                fecha=row.get("fecha", "").strip(),
            )
        )

    parsed = {
        "oid": oid,
        "nro_entrega": first.get("nro_entrega", "").strip(),
        "fecha": _parse_date(first.get("fecha", "")),
        "transporte_tipo": first.get("transporte_tipo", "").strip(),
        "flete": first.get("flete", "").strip(),
        "chofer": first.get("chofer", "").strip(),
        "acompanante": first.get("acompanante", "").strip(),
        "transporte": first.get("transporte", "").strip(),
        "remitos": remitos,
    }
    _validate_parsed_hoja(parsed)
    return parsed


def parse_csv_file(file_obj: Any) -> dict[str, Any]:
    if hasattr(file_obj, "seek"):
        try:
            file_obj.seek(0)
        except Exception:
            pass

    content = file_obj.read()
    if isinstance(content, bytes):
        content = content.decode("utf-8-sig")
    reader = csv.DictReader(content.splitlines())
    if not reader.fieldnames:
        raise ValueError("El CSV no contiene encabezados.")

    rows = []
    for row in reader:
        normalized = { _normalize_header(key): _normalize_value(value or "") for key, value in row.items() if key is not None }
        rows.append(normalized)

    if not rows:
        raise ValueError("El CSV no contiene filas de datos.")

    headers = set(rows[0].keys())
    missing = TABULAR_REQUIRED_COLUMNS - headers
    if missing:
        raise ValueError("Faltan columnas obligatorias: " + ", ".join(sorted(missing)))

    return _build_parsed_from_rows(rows)


def parse_xlsx_file(file_obj: Any) -> dict[str, Any]:
    if hasattr(file_obj, "seek"):
        try:
            file_obj.seek(0)
        except Exception:
            pass

    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(rows_iter)
    except StopIteration as exc:
        raise ValueError("El Excel no contiene datos.") from exc

    headers = [_normalize_header(str(header or "")) for header in raw_headers]
    missing = TABULAR_REQUIRED_COLUMNS - set(headers)
    if missing:
        raise ValueError("Faltan columnas obligatorias: " + ", ".join(sorted(missing)))

    rows = []
    for row in rows_iter:
        mapped = {}
        for index, header in enumerate(headers):
            raw = row[index] if index < len(row) else ""
            mapped[header] = _normalize_value(str(raw)) if raw is not None else ""
        if any(value for value in mapped.values()):
            rows.append(mapped)

    if not rows:
        raise ValueError("El Excel no contiene filas de datos.")

    return _build_parsed_from_rows(rows)


def import_tabular_file(file_obj: Any, original_file: Any) -> tuple[int, int]:
    parsed = parse_tabular_file(file_obj)

    if hasattr(original_file, "seek"):
        try:
            original_file.seek(0)
        except Exception:
            pass

    hoja = _import_parsed_hoja(parsed, original_file)
    return 1, hoja.remitos.count()


def parse_tabular_file(file_obj: Any) -> dict[str, Any]:
    name = getattr(file_obj, "name", "").lower()
    if name.endswith(".csv"):
        return parse_csv_file(file_obj)
    return parse_xlsx_file(file_obj)

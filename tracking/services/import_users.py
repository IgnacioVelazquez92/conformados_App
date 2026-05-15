from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from django.contrib.auth.models import User
from django.db import transaction
from openpyxl import load_workbook

from tracking.models import Empresa, RoleDefinition, UserProfile

USER_IMPORT_HEADERS = [
    "username",
    "nombre",
    "apellido",
    "email",
    "password",
    "rol",
    "empresa_principal",
    "empresas",
    "activo",
    "share_logistica",
    "share_cliente",
]
DEFAULT_IMPORT_PASSWORD = "Cambiar12345!"


@dataclass
class ImportedUserRow:
    username: str
    first_name: str
    last_name: str
    email: str
    password: str
    rol: str
    empresa_principal: Empresa
    empresas: list[Empresa]
    is_active: bool
    share_logistica: bool
    share_cliente: bool


@dataclass
class UserImportResult:
    created: int
    updated: int


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower()


def _normalize_value(value: Any) -> str:
    return str(value or "").strip() if value is not None else ""


def _parse_bool(value: Any, *, default: bool = False) -> bool:
    raw = _normalize_value(value).lower()
    if not raw:
        return default
    return raw in {"1", "si", "sí", "s", "true", "activo", "x"}


def _empresa_by_token(token: str) -> Empresa | None:
    value = token.strip()
    if not value:
        return None
    return Empresa.objects.filter(active=True).filter(code=value).first() or Empresa.objects.filter(
        active=True,
        slug=value,
    ).first()


def _parse_empresas(value: str, row_number: int) -> list[Empresa]:
    tokens = [token.strip() for token in value.replace(";", ",").split(",") if token.strip()]
    empresas: list[Empresa] = []
    seen: set[int] = set()
    for token in tokens:
        empresa = _empresa_by_token(token)
        if not empresa:
            raise ValueError(f"Fila {row_number}: empresa no encontrada o inactiva: {token}.")
        if empresa.pk not in seen:
            empresas.append(empresa)
            seen.add(empresa.pk)
    return empresas


def parse_users_excel(file_obj: Any) -> list[ImportedUserRow]:
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("El Excel no contiene datos.")

    headers = [_normalize_header(value) for value in rows[0]]
    required = {"username", "nombre", "apellido", "email", "rol", "empresas"}
    missing = required - set(headers)
    if missing:
        raise ValueError("Faltan columnas obligatorias: " + ", ".join(sorted(missing)))

    index = {header: position for position, header in enumerate(headers)}
    parsed_rows: list[ImportedUserRow] = []
    seen_usernames: set[str] = set()
    errors: list[str] = []

    for row_number, row in enumerate(rows[1:], start=2):
        values = {header: _normalize_value(row[position] if position < len(row) else "") for header, position in index.items()}
        if not any(values.values()):
            continue

        username = values.get("username", "").strip()
        if not username:
            errors.append(f"Fila {row_number}: falta username.")
            continue
        if username in seen_usernames:
            errors.append(f"Fila {row_number}: username duplicado en archivo: {username}.")
            continue
        seen_usernames.add(username)

        rol = values.get("rol", "").strip()
        if not RoleDefinition.objects.filter(code=rol, active=True).exists():
            errors.append(f"Fila {row_number}: rol no encontrado o inactivo: {rol}.")
            continue

        try:
            empresas = _parse_empresas(values.get("empresas", ""), row_number)
        except ValueError as exc:
            errors.append(str(exc))
            continue
        if not empresas:
            errors.append(f"Fila {row_number}: informa al menos una empresa en la columna empresas.")
            continue

        empresa_principal_token = values.get("empresa_principal", "").strip()
        empresa_principal = _empresa_by_token(empresa_principal_token) if empresa_principal_token else empresas[0]
        if not empresa_principal:
            errors.append(f"Fila {row_number}: empresa_principal no encontrada o inactiva: {empresa_principal_token}.")
            continue
        if empresa_principal.pk not in {empresa.pk for empresa in empresas}:
            errors.append(f"Fila {row_number}: empresa_principal debe estar incluida en empresas.")
            continue

        parsed_rows.append(
            ImportedUserRow(
                username=username,
                first_name=values.get("nombre", ""),
                last_name=values.get("apellido", ""),
                email=values.get("email", ""),
                password=values.get("password", ""),
                rol=rol,
                empresa_principal=empresa_principal,
                empresas=empresas,
                is_active=_parse_bool(values.get("activo", ""), default=True),
                share_logistica=_parse_bool(values.get("share_logistica", "")),
                share_cliente=_parse_bool(values.get("share_cliente", "")),
            )
        )

    if errors:
        raise ValueError("Importacion de usuarios fallida:\n- " + "\n- ".join(errors))
    if not parsed_rows:
        raise ValueError("El Excel no contiene usuarios para importar.")
    return parsed_rows


@transaction.atomic
def import_users_excel(file_obj: Any) -> UserImportResult:
    rows = parse_users_excel(file_obj)
    created = 0
    updated = 0

    for row in rows:
        user, was_created = User.objects.get_or_create(username=row.username)
        user.first_name = row.first_name
        user.last_name = row.last_name
        user.email = row.email
        user.is_active = row.is_active
        if row.password:
            user.set_password(row.password)
        elif was_created:
            user.set_password(DEFAULT_IMPORT_PASSWORD)
        user.save()

        profile, _ = UserProfile.objects.get_or_create(user=user)
        profile.rol = row.rol
        profile.empresa_principal = row.empresa_principal
        profile.share_logistica = row.share_logistica
        profile.share_cliente = row.share_cliente
        profile.save()
        profile.empresas.set(row.empresas)

        if was_created:
            created += 1
        else:
            updated += 1

    return UserImportResult(created=created, updated=updated)
